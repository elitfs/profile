from __future__ import annotations

import csv
import datetime as dt
import io
import json
import math
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parent
HOST = "127.0.0.1"
PORT = 4174
WORLD_BANK_API = "https://api.worldbank.org/v2"

COUNTRIES = {
    "NG": "Nigeria",
    "GH": "Ghana",
    "KE": "Kenya",
    "ZA": "South Africa",
    "US": "United States",
    "GB": "United Kingdom",
    "CA": "Canada",
    "IN": "India",
    "AE": "United Arab Emirates",
    "EUU": "European Union",
}

INDICATORS = {
    "inflation": "FP.CPI.TOTL.ZG",
    "gdpGrowth": "NY.GDP.MKTP.KD.ZG",
}

DEFAULT_SCENARIOS = [
    {"scenario": "Best Case", "parameter": 0.9, "weight": 0.3},
    {"scenario": "Base Case", "parameter": 1.0, "weight": 0.4},
    {"scenario": "Worse Case", "parameter": 1.1, "weight": 0.3},
]

DEFAULT_SCORECARD = [
    {"criterion": "LPO genuity", "weight": 0.05, "maxScore": 1},
    {"criterion": "Credit search", "weight": 0.10, "maxScore": 1},
    {"criterion": "Collateral as a % of loan amount", "weight": 0.30, "maxScore": 3},
    {"criterion": "Net cashflow as a % of repayment amount", "weight": 0.30, "maxScore": 3},
    {"criterion": "No of years in business", "weight": 0.05, "maxScore": 2},
    {"criterion": "Nature of business", "weight": 0.05, "maxScore": 5},
    {"criterion": "No of days past due", "weight": 0.10, "maxScore": 3},
    {"criterion": "Restructuring", "weight": 0.05, "maxScore": 1},
]

CLASSIFICATION_PROFILES = {
    "performing": {"basePd12": 0.020, "lgd": 0.45, "ccf": 0.00, "stage": 1},
    "standard": {"basePd12": 0.020, "lgd": 0.45, "ccf": 0.00, "stage": 1},
    "normal": {"basePd12": 0.020, "lgd": 0.45, "ccf": 0.00, "stage": 1},
    "current": {"basePd12": 0.018, "lgd": 0.42, "ccf": 0.00, "stage": 1},
    "pass": {"basePd12": 0.018, "lgd": 0.42, "ccf": 0.00, "stage": 1},
    "staff": {"basePd12": 0.012, "lgd": 0.30, "ccf": 0.00, "stage": 1},
    "lpo": {"basePd12": 0.050, "lgd": 0.45, "ccf": 0.20, "stage": 1},
    "trade": {"basePd12": 0.060, "lgd": 0.50, "ccf": 0.00, "stage": 2},
    "overdraft": {"basePd12": 0.070, "lgd": 0.55, "ccf": 0.75, "stage": 1},
    "watch": {"basePd12": 0.080, "lgd": 0.50, "ccf": 0.10, "stage": 2},
    "special mention": {"basePd12": 0.100, "lgd": 0.52, "ccf": 0.10, "stage": 2},
    "substandard": {"basePd12": 0.250, "lgd": 0.60, "ccf": 0.20, "stage": 2},
    "doubtful": {"basePd12": 0.500, "lgd": 0.75, "ccf": 0.50, "stage": 3},
    "non performing": {"basePd12": 1.000, "lgd": 0.80, "ccf": 0.50, "stage": 3},
    "npl": {"basePd12": 1.000, "lgd": 0.80, "ccf": 0.50, "stage": 3},
    "lost": {"basePd12": 1.000, "lgd": 0.95, "ccf": 0.50, "stage": 3},
    "default": {"basePd12": 1.000, "lgd": 0.85, "ccf": 0.50, "stage": 3},
}

COLLATERAL_HAIRCUTS = {
    "cash": 0.95,
    "deposit": 0.95,
    "treasury": 0.90,
    "government": 0.90,
    "bond": 0.80,
    "real estate": 0.70,
    "property": 0.70,
    "vehicle": 0.55,
    "equipment": 0.50,
    "inventory": 0.45,
    "receivable": 0.45,
    "guarantee": 0.35,
    "unsecured": 0.00,
}

FIELD_ALIASES = {
    "customerId": ["customer id", "customer_id", "client id", "obligor id", "account number", "account no", "asset id", "loan id", "facility id", "id"],
    "customerName": ["customer name", "borrower", "customer", "client", "counterparty", "obligor", "name"],
    "disbursementDate": ["disbursement date", "loan date", "origination date", "start date", "booking date"],
    "dueDate": ["due date", "maturity date", "expiry date", "end date", "repayment date"],
    "periodEnd": ["period end", "reporting date", "assessment date", "as of date", "valuation date"],
    "amount": ["amount", "loan amount", "principal", "facility amount", "approved amount", "limit", "total line"],
    "outstanding": ["outstanding", "outstanding balance", "outsanding bal", "outsanding loan balance", "gross carrying amount", "amortized cost", "amortised cost", "carrying amount", "ead", "balance"],
    "loanClassification": ["loan classification", "classification", "risk classification", "asset class", "product", "loan type", "facility type", "portfolio", "segment"],
    "daysPastDue": ["past due day", "past due days", "days past due", "dpd", "arrears days", "no. of days past due", "no of days past due"],
    "stage": ["stage", "ifrs 9 stage", "ifrs9 stage", "ecl stage"],
    "eir": ["eir", "effective interest rate", "effective interest rate %", "interest rate", "discount rate", "discount rate %"],
    "undrawnLimit": ["undrawn limit", "undrawn", "unused limit", "limit undrawn", "off balance sheet exposure"],
    "creditLimit": ["credit limit", "total line", "approved limit", "facility limit"],
    "ccf": ["ccf", "ccf %", "credit conversion factor", "credit conversion factor %"],
    "collateralValue": ["collateral value", "security value", "forced sale value", "fsv", "collateral", "collateral amount"],
    "collateralType": ["collateral type", "security type", "security", "collateral class"],
    "netCashflowCoverage": ["net cashflow as a % of repayment amount", "net cashflow coverage", "cashflow coverage", "cash flow coverage", "dscr"],
    "collateralCoverage": ["collateral as a % of loan amount", "collateral coverage", "security coverage", "collateral coverage %"],
    "yearsInBusiness": ["no of years in business", "years in business", "business age", "borrower age years"],
    "businessType": ["nature of business", "business type", "legal form", "company type"],
    "creditSearch": ["credit search", "credit bureau", "bureau status", "credit check"],
    "lpoGenuity": ["lpo genuity", "lpo genuineness", "lpo genuine", "contract genuity"],
    "restructuring": ["restructuring", "restructured", "forbearance", "modified"],
    "defaulted": ["defaulted", "default flag", "credit impaired", "impaired", "npl", "non performing", "default"],
    "pd12": ["12m pd", "12m pd %", "pd12", "pd12m", "pd 12m", "12 month pd"],
    "lifetimePd": ["lifetime pd", "lifetime pd %", "lt pd", "cumulative pd"],
    "lgd": ["lgd", "lgd %", "loss given default"],
}

SAMPLE_ROWS = [
    {
        "Customer ID": "CUST-001",
        "Customer Name": "Aster Manufacturing Ltd",
        "Disbursement Date": "2024-02-15",
        "Due Date": "2027-02-15",
        "Loan Amount": 125000000,
        "Outstanding Balance": 118000000,
        "Loan Classification": "Performing",
        "Past Due Days": 0,
        "EIR %": 13.0,
        "Collateral Type": "Real Estate",
        "Collateral Value": 160000000,
        "Years In Business": 8,
        "Credit Search": "Positive",
        "Restructuring": "No",
    },
    {
        "Customer ID": "CUST-002",
        "Customer Name": "Nile Trading Plc",
        "Disbursement Date": "2025-01-20",
        "Due Date": "2026-01-20",
        "Loan Amount": 38000000,
        "Outstanding Balance": 36000000,
        "Loan Classification": "Watchlist",
        "Past Due Days": 48,
        "EIR %": 11.0,
        "Collateral Type": "Receivable",
        "Collateral Value": 20000000,
        "Years In Business": 3,
        "Credit Search": "Positive",
        "Restructuring": "No",
    },
    {
        "Customer ID": "CUST-003",
        "Customer Name": "Kora Energy Services",
        "Disbursement Date": "2024-06-01",
        "Due Date": "2025-06-01",
        "Loan Amount": 73000000,
        "Outstanding Balance": 69000000,
        "Credit Limit": 90000000,
        "Loan Classification": "Overdraft",
        "Past Due Days": 97,
        "EIR %": 15.5,
        "Collateral Type": "Equipment",
        "Collateral Value": 35000000,
        "Years In Business": 1,
        "Credit Search": "Negative",
        "Restructuring": "Yes",
    },
]


class EclRequestHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def log_message(self, format, *args):
        sys.stdout.write("%s - - [%s] %s\n" % (self.address_string(), self.log_date_time_string(), format % args))

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            self.send_json({"ok": True, "app": "ifrs9-ecl"})
            return
        if parsed.path == "/api/countries":
            self.send_json({"countries": [{"code": code, "name": name} for code, name in COUNTRIES.items()]})
            return
        if parsed.path == "/api/macro":
            params = parse_qs(parsed.query)
            country = (params.get("country") or ["NG"])[0].upper()
            self.handle_macro(country)
            return
        if parsed.path == "/api/template":
            self.send_workbook(template_workbook(), "ifrs9-guided-ecl-template.xlsx")
            return
        super().do_GET()

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/compute":
            self.handle_compute()
            return
        if path == "/api/export":
            self.handle_export()
            return
        self.send_json({"error": "Endpoint not found."}, HTTPStatus.NOT_FOUND)

    def handle_macro(self, country):
        try:
            payload = fetch_macro_indicators(country)
            self.send_json(payload)
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_GATEWAY)

    def handle_compute(self):
        try:
            body = self.rfile.read(int(self.headers.get("Content-Length", "0")))
            files, fields = parse_multipart(body, self.headers.get("Content-Type", ""))
            main_file = files.get("file") or files.get("loanFile") or next(iter(files.values()), None)
            if not main_file:
                raise ValueError("No loan schedule was uploaded.")

            settings = json.loads(fields.get("settings") or "{}")
            workbook_sheets = workbook_rows_by_sheet(main_file["filename"], main_file["content"])
            schedule_name, schedule_rows = choose_schedule_sheet(workbook_sheets)
            if not schedule_rows:
                raise ValueError("No loan schedule rows were found in the uploaded file.")

            scorecard_rows = []
            scorecard_file = files.get("scorecard")
            if scorecard_file:
                _, scorecard_rows = choose_first_sheet(workbook_rows_by_sheet(scorecard_file["filename"], scorecard_file["content"]))
            else:
                scorecard_rows = find_optional_sheet(workbook_sheets, ["score card", "scorecard"]) or []

            scenario_rows = find_optional_sheet(workbook_sheets, ["weight", "scenario"]) or []
            scorecard = parse_scorecard(scorecard_rows)
            scenarios = parse_scenarios(scenario_rows) or DEFAULT_SCENARIOS
            mappings = detect_mappings(schedule_rows)
            results = [
                compute_guided_row(row, index, settings, mappings, scorecard, scenarios)
                for index, row in enumerate(schedule_rows)
            ]
            self.send_json(
                {
                    "rows": schedule_rows,
                    "mappings": mappings,
                    "results": results,
                    "summary": summarize(results),
                    "model": {
                        "scheduleSheet": schedule_name,
                        "scorecardRules": len(scorecard),
                        "scenarioMultiplier": scenario_multiplier(scenarios),
                    },
                }
            )
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def handle_export(self):
        try:
            body = self.rfile.read(int(self.headers.get("Content-Length", "0")))
            payload = json.loads(body.decode("utf-8") or "{}")
            rows = payload.get("rows") or []
            if not rows:
                raise ValueError("No rows were supplied for export.")
            workbook = workbook_from_rows(rows, "ECL Results")
            self.send_workbook(workbook, "ifrs9-ecl-results.xlsx")
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def send_json(self, payload, status=HTTPStatus.OK):
        data = json.dumps(payload, default=json_default).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_workbook(self, workbook, filename):
        output = io.BytesIO()
        workbook.save(output)
        data = output.getvalue()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)


def fetch_macro_indicators(country):
    country = country.upper()
    result = {
        "country": country,
        "countryName": COUNTRIES.get(country, country),
        "source": "World Bank World Development Indicators",
        "sourceUrl": "https://api.worldbank.org/v2",
        "indicators": {},
    }
    for key, indicator in INDICATORS.items():
        encoded = urllib.parse.quote(country)
        url = f"{WORLD_BANK_API}/country/{encoded}/indicator/{indicator}?format=json&per_page=10"
        with urllib.request.urlopen(url, timeout=12) as response:
            payload = json.loads(response.read().decode("utf-8"))
        observations = payload[1] if isinstance(payload, list) and len(payload) > 1 else []
        latest = next((item for item in observations if item.get("value") is not None), None)
        if not latest:
            raise ValueError(f"No World Bank data found for {indicator} and country {country}.")
        result["indicators"][key] = {
            "indicator": indicator,
            "name": latest.get("indicator", {}).get("value"),
            "year": latest.get("date"),
            "value": latest.get("value"),
        }
    return result


def parse_multipart(body: bytes, content_type: str):
    boundary_match = re.search(r'boundary="?([^";]+)"?', content_type)
    if not boundary_match:
        raise ValueError("Expected multipart form data.")
    boundary = ("--" + boundary_match.group(1)).encode("utf-8")
    files = {}
    fields = {}

    for raw_part in body.split(boundary):
        part = raw_part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        if part.endswith(b"--"):
            part = part[:-2].rstrip(b"\r\n")
        header_blob, separator, content = part.partition(b"\r\n\r\n")
        if not separator:
            continue
        headers = parse_part_headers(header_blob)
        disposition = headers.get("content-disposition", "")
        name = regex_group(r'name="([^"]+)"', disposition)
        filename = regex_group(r'filename="([^"]*)"', disposition)
        content = content.rstrip(b"\r\n")
        if filename:
            files[name or "file"] = {"name": name, "filename": Path(filename).name, "content": content}
        elif name:
            fields[name] = content.decode("utf-8", errors="replace")

    return files, fields


def parse_part_headers(header_blob: bytes):
    headers = {}
    for line in header_blob.decode("utf-8", errors="replace").split("\r\n"):
        if ":" in line:
            key, value = line.split(":", 1)
            headers[key.strip().lower()] = value.strip()
    return headers


def regex_group(pattern: str, value: str):
    match = re.search(pattern, value)
    return match.group(1) if match else None


def workbook_rows_by_sheet(filename: str, content: bytes):
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        text = content.decode("utf-8-sig")
        return {"Loan Schedule": [dict(row) for row in csv.DictReader(io.StringIO(text))]}
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        return {worksheet.title: rows_from_worksheet(worksheet) for worksheet in workbook.worksheets}
    if extension == ".xls":
        raise ValueError("Legacy .xls files are not supported. Save the workbook as .xlsx and upload again.")
    raise ValueError("Unsupported file type. Upload .xlsx, .xlsm, or .csv.")


def rows_from_worksheet(worksheet):
    raw_rows = list(worksheet.iter_rows(values_only=True))
    while raw_rows and all(is_blank(cell) for cell in raw_rows[0]):
        raw_rows.pop(0)
    if not raw_rows:
        return []

    header_index = choose_header_index(raw_rows)
    headers = [
        str(cell).strip() if not is_blank(cell) else f"Column {index + 1}"
        for index, cell in enumerate(raw_rows[header_index])
    ]
    records = []
    for values in raw_rows[header_index + 1 :]:
        if all(is_blank(cell) for cell in values):
            continue
        record = {}
        for index, header in enumerate(headers):
            record[header] = json_ready(values[index] if index < len(values) else "")
        records.append(record)
    return records


def choose_header_index(raw_rows):
    best_index = 0
    best_score = -1
    header_terms = set()
    for aliases in FIELD_ALIASES.values():
        header_terms.update(normalize_header(alias) for alias in aliases)
    header_terms.update(["criterion", "factor", "weight", "max score", "score", "scenario", "parameter"])
    for index, row in enumerate(raw_rows[:10]):
        values = [normalize_header(cell) for cell in row if not is_blank(cell)]
        score = sum(1 for value in values if value in header_terms or any(term in value for term in header_terms))
        score += min(len(values), 8) * 0.1
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def choose_schedule_sheet(workbook_sheets):
    preferred_terms = ["loan schedule", "schedule", "portfolio", "loan data", "input", "loans"]
    excluded_terms = ["score", "weight", "scenario", "macro", "summary", "lgd", "default", "assess"]
    candidates = []
    for name, rows in workbook_sheets.items():
        normalized = normalize_header(name)
        if not rows:
            continue
        score = len(rows)
        if any(term in normalized for term in preferred_terms):
            score += 1000
        if any(term in normalized for term in excluded_terms):
            score -= 1000
        candidates.append((score, name, rows))
    if not candidates:
        return "", []
    _, name, rows = max(candidates, key=lambda item: item[0])
    return name, rows


def choose_first_sheet(workbook_sheets):
    for name, rows in workbook_sheets.items():
        if rows:
            return name, rows
    return "", []


def find_optional_sheet(workbook_sheets, name_terms):
    for name, rows in workbook_sheets.items():
        normalized = normalize_header(name)
        if any(term in normalized for term in name_terms):
            return rows
    return []


def detect_mappings(rows):
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = [{"raw": header, "normalized": normalize_header(header)} for header in headers]
    mappings = {}
    for field, aliases in FIELD_ALIASES.items():
        normalized_aliases = [normalize_header(alias) for alias in aliases]
        match = next((header for header in normalized_headers if header["normalized"] in normalized_aliases), None)
        if not match:
            match = next(
                (
                    header
                    for header in normalized_headers
                    if any(alias in header["normalized"] for alias in normalized_aliases)
                ),
                None,
            )
        if match:
            mappings[field] = match["raw"]
    return mappings


def parse_scorecard(rows):
    if not rows:
        return DEFAULT_SCORECARD
    parsed = []
    mappings = detect_scorecard_mappings(rows)
    for row in rows:
        criterion = clean_text(get_by_mapping(row, mappings, "criterion"))
        if not criterion:
            continue
        weight = percent_or_number(get_by_mapping(row, mappings, "weight"), None)
        max_score = numeric_input(get_by_mapping(row, mappings, "maxScore"), None)
        if weight is None or max_score is None or max_score <= 0:
            continue
        parsed.append({"criterion": criterion, "weight": weight, "maxScore": max_score})
    return parsed or DEFAULT_SCORECARD


def detect_scorecard_mappings(rows):
    aliases = {
        "criterion": ["criterion", "factor", "lpo financing", "score factor", "risk factor", "parameter"],
        "weight": ["weight", "score weight", "weighted", "d"],
        "maxScore": ["max score", "maximum score", "max", "score max"],
    }
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = [{"raw": header, "normalized": normalize_header(header)} for header in headers]
    mappings = {}
    for field, field_aliases in aliases.items():
        normalized_aliases = [normalize_header(alias) for alias in field_aliases]
        match = next(
            (
                header
                for header in normalized_headers
                if header["normalized"] in normalized_aliases
                or any(alias in header["normalized"] or header["normalized"] in alias for alias in normalized_aliases)
            ),
            None,
        )
        if match:
            mappings[field] = match["raw"]
    return mappings


def parse_scenarios(rows):
    if not rows:
        return []
    mappings = detect_generic_mappings(
        rows,
        {
            "scenario": ["scenario", "scenerio", "case"],
            "parameter": ["parameter", "multiplier", "factor"],
            "weight": ["probability weight", "weight", "probability"],
        },
    )
    scenarios = []
    for row in rows:
        scenario = clean_text(get_by_mapping(row, mappings, "scenario"))
        if not scenario or scenario.lower() == "total":
            continue
        parameter = percent_or_number(get_by_mapping(row, mappings, "parameter"), None)
        weight = percent_or_number(get_by_mapping(row, mappings, "weight"), None)
        if parameter is not None and weight is not None:
            scenarios.append({"scenario": scenario, "parameter": parameter, "weight": weight})
    return scenarios


def detect_generic_mappings(rows, alias_map):
    headers = list(rows[0].keys()) if rows else []
    normalized_headers = [{"raw": header, "normalized": normalize_header(header)} for header in headers]
    mappings = {}
    for field, aliases in alias_map.items():
        normalized_aliases = [normalize_header(alias) for alias in aliases]
        match = next(
            (
                header
                for header in normalized_headers
                if header["normalized"] in normalized_aliases
                or any(alias in header["normalized"] or header["normalized"] in alias for alias in normalized_aliases)
            ),
            None,
        )
        if match:
            mappings[field] = match["raw"]
    return mappings


def compute_guided_row(row, index, settings, mappings, scorecard, scenarios):
    def value(field):
        return get_by_mapping(row, mappings, field)

    customer_id = clean_text(value("customerId")) or f"Row {index + 1}"
    customer_name = clean_text(value("customerName")) or "Unassigned"
    loan_class = clean_text(value("loanClassification")) or "Performing"
    profile = classification_profile(loan_class)
    amount = numeric_input(value("amount"), 0)
    outstanding = numeric_input(value("outstanding"), amount)
    eir = percent_value(value("eir"), percent_value(settings.get("defaultDiscount"), 0.12))
    days_past_due = numeric_input(value("daysPastDue"), 0)
    due_date = parse_date(value("dueDate"))
    period_end = parse_date(value("periodEnd")) or dt.date.today()
    if due_date and days_past_due == 0 and due_date < period_end:
        days_past_due = (period_end - due_date).days

    stage, stage_rationale = determine_stage(row, mappings, loan_class, days_past_due, settings, profile)
    remaining_months = remaining_life_months(due_date, period_end)
    ead, ead_rationale = compute_ead(row, mappings, amount, outstanding, profile)
    score_result = compute_score(row, mappings, scorecard, days_past_due)
    macro = macro_settings(settings)
    macro_multiplier = compute_macro_multiplier(macro["inflationRate"], macro["gdpGrowthRate"])
    scenario_factor = scenario_multiplier(scenarios)

    uploaded_pd12 = percent_value(value("pd12"), None)
    uploaded_lifetime_pd = percent_value(value("lifetimePd"), None)
    if stage == 3 or truthy(value("defaulted")):
        pd12 = 1.0
        lifetime_pd = 1.0
        pd_basis = "Default or Stage 3: PD set to 100%"
    elif uploaded_pd12 is not None:
        pd12 = uploaded_pd12
        lifetime_pd = uploaded_lifetime_pd if uploaded_lifetime_pd is not None else annual_to_lifetime_pd(pd12, remaining_months)
        pd_basis = "Uploaded PD used"
    else:
        base_pd = profile["basePd12"]
        score_multiplier = 1 + ((1 - score_result["scoreRatio"]) * 2.0)
        dpd_multiplier = dpd_pd_multiplier(days_past_due)
        pd12 = clamp(base_pd * score_multiplier * dpd_multiplier * macro_multiplier, 0.001, 0.95)
        lifetime_pd = annual_to_lifetime_pd(pd12, remaining_months)
        pd_basis = "Classification + scorecard + DPD + macro"

    uploaded_lgd = percent_value(value("lgd"), None)
    if uploaded_lgd is not None:
        lgd = uploaded_lgd
        lgd_basis = "Uploaded LGD used"
    else:
        lgd, lgd_basis = compute_lgd(row, mappings, ead, eir, profile, macro_multiplier)

    horizon_months = min(12, remaining_months) if stage == 1 else remaining_months
    pd_used = pd12 if stage == 1 else lifetime_pd
    discount_factor = 1 / math.pow(1 + eir, horizon_months / 12) if eir > 0 else 1
    ecl = ead * pd_used * lgd * discount_factor * scenario_factor

    return {
        "assetId": customer_id,
        "customerId": customer_id,
        "borrower": customer_name,
        "customerName": customer_name,
        "assetClass": loan_class,
        "loanClassification": loan_class,
        "stage": stage,
        "stageLabel": f"Stage {stage}",
        "rationale": stage_rationale,
        "ead": ead,
        "eadBasis": ead_rationale,
        "exposureBase": outstanding,
        "pd12": pd12,
        "lifetimePd": lifetime_pd,
        "pdUsed": pd_used,
        "pdBasis": pd_basis,
        "score": score_result["score"],
        "scoreRatio": score_result["scoreRatio"],
        "scoreBasis": score_result["basis"],
        "lgd": lgd,
        "lgdBasis": lgd_basis,
        "discountRate": eir,
        "horizonMonths": horizon_months,
        "remainingLifeMonths": remaining_months,
        "discountFactor": discount_factor,
        "scenarioWeight": scenario_factor,
        "macroMultiplier": macro_multiplier,
        "inflationRate": macro["inflationRate"],
        "gdpGrowthRate": macro["gdpGrowthRate"],
        "ecl": ecl,
        "daysPastDue": days_past_due,
        "warnings": "",
    }


def determine_stage(row, mappings, loan_class, days_past_due, settings, profile):
    explicit_stage = parse_stage(get_by_mapping(row, mappings, "stage"))
    defaulted = truthy(get_by_mapping(row, mappings, "defaulted"))
    restructured = truthy(get_by_mapping(row, mappings, "restructuring"))
    stage2_dpd = numeric_input(settings.get("stage2Dpd"), 30)
    stage3_dpd = numeric_input(settings.get("stage3Dpd"), 90)
    normalized_class = normalize_header(loan_class)

    if explicit_stage:
        return explicit_stage, "Uploaded IFRS 9 stage retained"
    if defaulted or days_past_due > stage3_dpd or profile.get("stage") == 3:
        return 3, "Default, credit-impaired, Stage 3 classification, or over 90 DPD"
    if restructured or days_past_due > stage2_dpd or profile.get("stage") == 2:
        return 2, "SICR trigger: restructuring, classification, or over 30 DPD"
    if any(term in normalized_class for term in ["trade", "lease", "contract", "receivable"]):
        return 2, "Simplified lifetime ECL approach for receivable-like assets"
    return 1, "Performing asset with no SICR/default trigger"


def classification_profile(loan_class):
    normalized = normalize_header(loan_class)
    for key, profile in CLASSIFICATION_PROFILES.items():
        if key in normalized:
            return profile
    return {"basePd12": 0.035, "lgd": 0.50, "ccf": 0.00, "stage": 1}


def compute_ead(row, mappings, amount, outstanding, profile):
    credit_limit = numeric_input(get_by_mapping(row, mappings, "creditLimit"), amount)
    undrawn = numeric_input(get_by_mapping(row, mappings, "undrawnLimit"), max(credit_limit - outstanding, 0))
    ccf = percent_value(get_by_mapping(row, mappings, "ccf"), profile.get("ccf", 0))
    ead = max(outstanding + (undrawn * ccf), 0)
    if undrawn > 0 and ccf > 0:
        return ead, "Outstanding balance plus CCF-adjusted undrawn exposure"
    return ead, "Outstanding/amortised cost balance"


def compute_score(row, mappings, scorecard, days_past_due):
    total_weight = 0
    weighted_score = 0
    basis = []
    for rule in scorecard:
        criterion = rule["criterion"]
        weight = rule["weight"]
        max_score = rule["maxScore"]
        raw_score = criterion_score(criterion, row, mappings, days_past_due, max_score)
        if raw_score is None:
            continue
        total_weight += weight
        weighted_score += clamp(raw_score / max_score, 0, 1) * weight
        basis.append(f"{criterion}: {raw_score:g}/{max_score:g}")

    if total_weight <= 0:
        fallback_ratio = clamp(1 - (days_past_due / 120), 0, 1)
        return {"score": fallback_ratio, "scoreRatio": fallback_ratio, "basis": "Fallback behavioural score from DPD"}

    score_ratio = clamp(weighted_score / total_weight, 0, 1)
    return {"score": weighted_score, "scoreRatio": score_ratio, "basis": "; ".join(basis[:5])}


def criterion_score(criterion, row, mappings, days_past_due, max_score):
    text = normalize_header(criterion)
    if "lpo" in text:
        value = get_by_mapping(row, mappings, "lpoGenuity")
        return score_yes_no(value, max_score, yes_score=max_score, no_score=0)
    if "credit search" in text or "credit bureau" in text:
        value = clean_text(get_by_mapping(row, mappings, "creditSearch")).lower()
        if not value:
            return None
        return max_score if any(term in value for term in ["positive", "clean", "good", "pass", "yes"]) else 0
    if "collateral" in text and "%" in text or "collateral as" in text:
        coverage = percent_or_number(get_by_mapping(row, mappings, "collateralCoverage"), None)
        if coverage is None:
            amount = numeric_input(get_by_mapping(row, mappings, "amount"), 0)
            collateral = numeric_input(get_by_mapping(row, mappings, "collateralValue"), 0)
            coverage = collateral / amount if amount else None
        return coverage_score(coverage, max_score)
    if "cashflow" in text or "cash flow" in text:
        coverage = percent_or_number(get_by_mapping(row, mappings, "netCashflowCoverage"), None)
        return coverage_score(coverage, max_score)
    if "years in business" in text:
        years = numeric_input(get_by_mapping(row, mappings, "yearsInBusiness"), None)
        if years is None:
            return None
        if years > 5:
            return min(max_score, 2)
        if years >= 1:
            return min(max_score, 1)
        return 0
    if "nature of business" in text or "business type" in text:
        business_type = normalize_header(get_by_mapping(row, mappings, "businessType"))
        if not business_type:
            return None
        if "plc" in business_type:
            return min(max_score, 5)
        if "ltd" in business_type or "limited" in business_type:
            return min(max_score, 3)
        if "sole" in business_type:
            return min(max_score, 1)
        return min(max_score, 2)
    if "past due" in text or "dpd" in text:
        if days_past_due <= 30:
            return min(max_score, 3)
        if days_past_due <= 60:
            return min(max_score, 2)
        if days_past_due <= 90:
            return min(max_score, 1)
        return 0
    if "restructur" in text or "forbear" in text:
        value = get_by_mapping(row, mappings, "restructuring")
        return score_yes_no(value, max_score, yes_score=0, no_score=max_score)
    return None


def coverage_score(coverage, max_score):
    if coverage is None:
        return None
    if coverage > 10:
        coverage = coverage / 100
    if coverage >= 5:
        return min(max_score, 3)
    if coverage >= 3:
        return min(max_score, 2)
    if coverage >= 1:
        return min(max_score, 1)
    return 0


def score_yes_no(value, max_score, yes_score, no_score):
    text = clean_text(value).lower()
    if not text:
        return None
    if text in {"yes", "y", "true", "1", "positive", "good"}:
        return yes_score
    if text in {"no", "n", "false", "0", "negative", "bad"}:
        return no_score
    return None


def dpd_pd_multiplier(days_past_due):
    if days_past_due <= 0:
        return 0.8
    if days_past_due <= 30:
        return 1.0
    if days_past_due <= 60:
        return 2.0
    if days_past_due <= 90:
        return 4.0
    return 12.0


def macro_settings(settings):
    return {
        "inflationRate": percent_value(settings.get("inflationRate"), 0.10),
        "gdpGrowthRate": percent_value(settings.get("gdpGrowthRate"), 0.03),
    }


def compute_macro_multiplier(inflation_rate, gdp_growth_rate):
    inflation_pct = inflation_rate * 100
    gdp_pct = gdp_growth_rate * 100
    multiplier = 1 + ((inflation_pct - 10) * 0.025) - ((gdp_pct - 3) * 0.04)
    return clamp(multiplier, 0.70, 1.80)


def annual_to_lifetime_pd(pd12, remaining_months):
    years = max(remaining_months / 12, 1 / 12)
    return clamp(1 - math.pow(1 - clamp(pd12, 0, 1), years), 0, 1)


def compute_lgd(row, mappings, ead, eir, profile, macro_multiplier):
    if ead <= 0:
        return 0, "No exposure"
    recoveries = recovery_cashflows(row)
    if any(value > 0 for value in recoveries):
        pv = sum(value / math.pow(1 + (eir / 12), index + 1) for index, value in enumerate(recoveries))
        return clamp(1 - min(pv, ead) / ead, 0.02, 1), "Discounted expected recovery cash flows"

    collateral_value = numeric_input(get_by_mapping(row, mappings, "collateralValue"), 0)
    if collateral_value > 0:
        collateral_type = clean_text(get_by_mapping(row, mappings, "collateralType"))
        haircut = collateral_haircut(collateral_type)
        recovery = collateral_value * haircut
        lgd = 1 - min(recovery, ead) / ead
        return clamp(lgd * (0.85 + (macro_multiplier * 0.15)), 0.02, 1), f"Collateral value with {haircut:.0%} recovery haircut"

    base_lgd = profile.get("lgd", 0.50)
    return clamp(base_lgd * (0.85 + (macro_multiplier * 0.15)), 0.05, 1), "Default LGD by loan classification"


def recovery_cashflows(row):
    recoveries = []
    normalized = {normalize_header(key): value for key, value in row.items()}
    for month in range(1, 13):
        candidates = [
            f"month {month}",
            f"recovery month {month}",
            f"recovery {month}",
            f"cashflow month {month}",
            f"cash flow month {month}",
        ]
        value = next((normalized[candidate] for candidate in candidates if candidate in normalized), 0)
        recoveries.append(numeric_input(value, 0))
    return recoveries


def collateral_haircut(collateral_type):
    normalized = normalize_header(collateral_type)
    for key, haircut in COLLATERAL_HAIRCUTS.items():
        if key in normalized:
            return haircut
    return 0.40 if normalized else 0.00


def scenario_multiplier(scenarios):
    total_weight = sum(scenario.get("weight", 0) for scenario in scenarios)
    if total_weight <= 0:
        return 1
    return sum(scenario.get("parameter", 1) * scenario.get("weight", 0) for scenario in scenarios) / total_weight


def remaining_life_months(due_date, period_end):
    if not due_date:
        return 12
    months = (due_date.year - period_end.year) * 12 + (due_date.month - period_end.month)
    if due_date.day > period_end.day:
        months += 1
    return max(months, 1)


def template_workbook():
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "Loan Schedule"
    write_rows(schedule, SAMPLE_ROWS)

    scorecard = workbook.create_sheet("Scorecard")
    write_rows(
        scorecard,
        [
            {"Criterion": row["criterion"], "Weight": row["weight"], "Max Score": row["maxScore"]}
            for row in DEFAULT_SCORECARD
        ],
    )

    weights = workbook.create_sheet("Scenario Weights")
    write_rows(
        weights,
        [
            {"Scenario": row["scenario"], "Parameter": row["parameter"], "Probability Weight": row["weight"]}
            for row in DEFAULT_SCENARIOS
        ],
    )

    guidance = workbook.create_sheet("Guidance")
    guidance.append(["Purpose", "Required or optional", "Notes"])
    guidance.append(["Customer ID / Name", "Required", "Used to identify each loan."])
    guidance.append(["Due Date and Past Due Days", "Required", "Used for Stage 1, Stage 2, and Stage 3 assessment."])
    guidance.append(["Outstanding Balance", "Required", "Used as EAD/EOD. If blank, Loan Amount is used."])
    guidance.append(["Loan Classification", "Recommended", "Used for base PD, LGD, CCF, and stage hints."])
    guidance.append(["Collateral Value / Type", "Optional", "Used to estimate LGD when recovery cash flows are not supplied."])
    guidance.append(["Month 1 to Month 12 recoveries", "Optional", "Used to discount expected recoveries for LGD."])
    for worksheet in workbook.worksheets:
        autofit(worksheet)
    return workbook


def workbook_from_rows(rows, sheet_name):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    write_rows(worksheet, rows)
    autofit(worksheet)
    return workbook


def write_rows(worksheet, rows):
    if not rows:
        return
    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])


def autofit(worksheet):
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 44)


def get_by_mapping(row, mappings, field):
    mapped = mappings.get(field)
    return row.get(mapped, "") if mapped else ""


def percent_value(value, fallback):
    if value is None or clean_text(value) == "":
        return fallback
    raw = numeric_input(value, None)
    if raw is None:
        return fallback
    return clamp(raw / 100 if raw > 1 else raw, 0, 1)


def percent_or_number(value, fallback):
    if value is None or clean_text(value) == "":
        return fallback
    raw = numeric_input(value, None)
    if raw is None:
        return fallback
    return raw / 100 if raw > 1 and "%" in clean_text(value) else raw


def numeric_input(value, fallback=0):
    if value is None:
        return fallback
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return value if math.isfinite(value) else fallback
    cleaned = clean_text(value).replace(",", "").replace("%", "")
    if not cleaned:
        return fallback
    try:
        number = float(cleaned)
        return number if math.isfinite(number) else fallback
    except ValueError:
        return fallback


def parse_stage(value):
    match = re.search(r"[123]", clean_text(value).lower())
    return int(match.group(0)) if match else None


def parse_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    try:
        return dt.datetime.fromisoformat(text).date()
    except ValueError:
        return None


def truthy(value):
    return clean_text(value).lower() in {
        "yes",
        "y",
        "true",
        "1",
        "default",
        "defaulted",
        "impaired",
        "npl",
        "non performing",
        "stage 3",
        "sicr",
    }


def normalize_header(value):
    return re.sub(r"\s+", " ", re.sub(r"[_-]+", " ", re.sub(r"[%()]", "", clean_text(value).lower()))).strip()


def clean_text(value):
    return str(value or "").strip()


def clamp(value, lower, upper):
    return min(max(value, lower), upper)


def is_blank(value):
    return value is None or str(value).strip() == ""


def json_ready(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    if value is None:
        return ""
    return value


def summarize(results):
    total_exposure = sum(row["ead"] for row in results)
    total_ecl = sum(row["ecl"] for row in results)
    return {
        "totalExposure": total_exposure,
        "totalEcl": total_ecl,
        "coverageRatio": total_ecl / total_exposure if total_exposure else 0,
        "rowsProcessed": len(results),
        "stage1": sum(1 for row in results if row["stage"] == 1),
        "stage2": sum(1 for row in results if row["stage"] == 2),
        "stage3": sum(1 for row in results if row["stage"] == 3),
    }


def json_default(value):
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    raise TypeError(f"Object of type {type(value).__name__} is not JSON serializable")


def main():
    server = ThreadingHTTPServer((HOST, PORT), EclRequestHandler)
    print(f"IFRS 9 ECL app running at http://{HOST}:{PORT}/")
    server.serve_forever()


if __name__ == "__main__":
    main()
